import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import shutil
import tkinter as tk
from tkinter import simpledialog

def create_version_list(folder_path, output_path):
    # "UI"を含むフォルダのみを取得
    folder_list = [f for f in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, f)) and "UI" in f]
    
    # データを準備
    data = {
        "フォルダ名": folder_list,
        "更新日時": [],
        "内容": []
    }

    for folder in folder_list:
        # フォルダの更新日時を取得
        folder_path_full = os.path.join(folder_path, folder)
        update_time = datetime.fromtimestamp(os.path.getmtime(folder_path_full)).strftime("%Y/%m/%d %H:%M:%S")
        data["更新日時"].append(update_time)

        # 初期値として内容を空欄に設定
        data["内容"].append("")

    # DataFrameを作成
    df = pd.DataFrame(data)

    # Excelファイルに出力
    try:
        df.to_excel(output_path, index=False, sheet_name="バージョン情報")

        # スタイルを設定
        wb = load_workbook(output_path)
        ws = wb.active

        # ヘッダーのスタイル
        header_fill = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FF000000")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill
            cell.border = border

        # データのスタイル
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

        # 列幅の調整
        ws.column_dimensions['A'].width = 30  # フォルダ名
        ws.column_dimensions['B'].width = 20  # 更新日時
        ws.column_dimensions['C'].width = 50  # 内容

        # 行の高さを調整
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 25

        wb.save(output_path)
        print(f"バージョン情報一覧を {output_path} に保存しました！")
    except PermissionError:
        print("Excelファイルが開いているため、保存できません。")
    except Exception as e:
        print(f"エラーが発生しました: {e}")

def copy_ui_folder(source_folder, destination_folder):
    try:
        shutil.copytree(source_folder, destination_folder)
        print(f"{source_folder} を {destination_folder} にコピーしました。")
    except FileExistsError:
        print(f"{destination_folder} は既に存在します。")
    except Exception as e:
        print(f"コピー中にエラーが発生しました: {e}")

# 使用例
folder_path = "C:\\Users\\NishiharaKatsuhiko\\Desktop\\python"  # フォルダのパスを指定
output_path = "C:\\Users\\NishiharaKatsuhiko\\Desktop\\python\\バージョン情報一覧.xlsx"  # 出力するExcelファイルのフルパスを指定

# UIフォルダのコピー
source_folder = os.path.join(folder_path, "UI")

# ダイアログボックスでフォルダ名を入力
root = tk.Tk()
root.withdraw()  # メインウィンドウを表示しない
destination_folder_name = simpledialog.askstring("フォルダ名入力", "コピー先のフォルダ名を入力してください:")
if destination_folder_name:
    destination_folder = os.path.join(folder_path, destination_folder_name)
    copy_ui_folder(source_folder, destination_folder)

    # バージョン一覧を作成
    create_version_list(folder_path, output_path)
else:
    print("フォルダ名が入力されませんでした。")
